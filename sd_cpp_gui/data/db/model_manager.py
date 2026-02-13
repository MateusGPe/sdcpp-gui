"""
Model Manager
"""

import csv
import json
import os
import time
import uuid
from typing import Any, Dict, List, Optional, Union, cast

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.worksheet import Worksheet

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    from typing import Any as Worksheet
    from typing import Any as openpyxl

from sd_cpp_gui.data.db.base_manager import ImportExportMixin
from sd_cpp_gui.data.db.database import db
from sd_cpp_gui.data.db.init_db import Database
from sd_cpp_gui.data.db.models import ModelData, ModelEntry
from sd_cpp_gui.data.remote.types import RemoteVersionDTO
from sd_cpp_gui.infrastructure.logger import get_logger

logger = get_logger(__name__)


class ModelManager(ImportExportMixin):
    """CRUD for Models."""

    def __init__(self) -> None:
        """Logic: Initializes DB."""
        Database()

    def register_from_remote(
        self, file_path: str, metadata: RemoteVersionDTO
    ) -> None:
        """
        Registers or updates a local model file using metadata from
        a remote source.

        Args:
                file_path: Absolute path to the local file.
                metadata: DTO containing remote metadata (Civitai, etc).
        """
        file_path = os.path.abspath(file_path)
        name = metadata.get("name", "Unknown Model")
        with db.atomic():
            existing = ModelEntry.get_or_none(ModelEntry.path == file_path)
            data = {
                "name": name,
                "path": file_path,
                "remote_source": "civitai",
                "remote_id": str(metadata["model_id"]),
                "remote_version_id": str(metadata["id"]),
                "base_model": metadata.get("base_model"),
                "description": metadata.get("description"),
            }
            if existing:
                ModelEntry.update(**data).where(
                    ModelEntry.id == existing.id
                ).execute()
            else:
                data["id"] = str(uuid.uuid4())
                data["params"] = "[]"
                ModelEntry.create(**data)

    def get_remote_index(self) -> Dict[str, str]:
        """Map remote_version_id -> path

        Logic: Returns map of remote IDs to local paths."""
        query = ModelEntry.select(ModelEntry.remote_version_id, ModelEntry.path)
        return {
            str(e.remote_version_id): str(e.path)
            for e in query
            if e.remote_version_id
        }

    def add_or_update_model(
        self,
        model_id: Optional[str],
        name: str,
        path: str,
        params: List[Dict[str, Any]],
        base_model: Optional[str] = None,
    ) -> None:
        """
        Manually adds or updates a model entry in the database.

        Args:
                model_id: Unique ID. If None, a timestamp-based ID is generated.
                name: Display name for the model.
                path: File system path to the model.
                params: List of default parameters for this model.
                base_model: Optional base model architecture (e.g., 'SD 1.5').
        """
        params_json = json.dumps(params)
        if not model_id:
            model_id = str(int(time.time()))
        data = {
            "id": model_id,
            "name": name,
            "path": path,
            "params": params_json,
        }
        if base_model is not None:
            data["base_model"] = base_model
        ModelEntry.replace(**data).execute()

    def get_model(self, model_id: str) -> Optional[ModelData]:
        """
        Retrieves a model's data by its ID.

        Args:
                model_id: The unique identifier.

        Returns:
                ModelData dictionary or None if not found.
        """
        entry = ModelEntry.get_or_none(ModelEntry.id == model_id)
        return self._entry_to_dict(entry) if entry else None

    def get_all(self) -> List[ModelData]:
        """
        Returns a list of all registered models, sorted by name.

        Returns:
                List of ModelData dictionaries.
        """
        query = ModelEntry.select().order_by(ModelEntry.name.asc())
        return [self._entry_to_dict(entry) for entry in query]

    def delete_model(self, model_id: str) -> None:
        """
        Removes a model entry from the database.

        Args:
                model_id: The ID of the model to delete.
        """
        ModelEntry.delete().where(ModelEntry.id == model_id).execute()

    def _entry_to_dict(self, entry: ModelEntry) -> ModelData:
        """Logic: Converts DB entry to TypedDict."""
        return ModelData(
            id=str(entry.id),
            name=str(entry.name),
            path=str(entry.path),
            params=self._safe_load_json(entry.params, []),
            remote_source=str(entry.remote_source),
            remote_id=str(entry.remote_id),
            remote_version_id=str(entry.remote_version_id),
            base_model=str(entry.base_model),
            description=str(entry.description),
        )

    def export_to_csv(self, filepath: str) -> None:
        """Exports models to CSV.

        Logic: Exports models to CSV."""
        data = self.get_all()
        if not data:
            return
        fieldnames = ["id", "name", "path", "params"]
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in data:
                row_copy = cast(Dict[str, Any], row.copy())
                filtered = {
                    k: v for k, v in row_copy.items() if k in fieldnames
                }
                filtered["params"] = json.dumps(
                    row["params"], ensure_ascii=False
                )
                writer.writerow(filtered)

    def export_to_xlsx(self, filepath: str) -> None:
        """Exports models to Excel (.xlsx).

        Logic: Exports models to Excel."""
        if not HAS_OPENPYXL:
            raise ImportError(
                "The 'openpyxl' library is required."
                " Install with: pip install openpyxl"
            )
        data = self.get_all()
        wb = openpyxl.Workbook()
        ws: Worksheet = cast(Worksheet, wb.active)  # type: ignore

        if not ws:
            return
        ws.title = "Models"
        headers = ["id", "name", "path", "params"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in data:
            params_str = json.dumps(row["params"], indent=2, ensure_ascii=False)
            if len(params_str) > 32000:
                params_str = params_str[:32000] + "... (truncated)"
            ws.append([row["id"], row["name"], row["path"], params_str])
            ws.cell(row=ws.max_row, column=4).alignment = Alignment(
                wrap_text=True
            )
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 50
        wb.save(filepath)

    def export_to_toml(self, filepath: str, root_key: str = "models") -> None:
        """Exports all models to a TOML file.

        Logic: Exports models to TOML."""
        super().export_to_toml(filepath, root_key=root_key)

    def import_from_toml(self, filepath: str, root_key: str = "models") -> None:
        """Imports models from a TOML file.

        Logic: Imports models from TOML."""
        super().import_from_toml(filepath, root_key=root_key)

    def import_from_csv(self, filepath: str) -> None:
        """Imports models from a CSV.

        Logic: Imports models from CSV."""
        with open(filepath, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            data: List[Dict[str, Any]] = []
            for row in reader:
                try:
                    row["params"] = json.loads(row["params"])  # type: ignore
                except (json.JSONDecodeError, TypeError):
                    row["params"] = []  # type: ignore
                data.append(row)  # type: ignore
            self._process_import_data(data)

    def import_from_xlsx(self, filepath: str) -> None:
        """Imports models from an Excel file (.xlsx).

        Logic: Imports models from Excel."""
        if not HAS_OPENPYXL:
            raise ImportError("Install 'openpyxl' for Excel support.")
        wb = openpyxl.load_workbook(filepath)
        ws: Worksheet = cast(Worksheet, wb.active)  # type: ignore

        if not ws:
            return
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(h).lower() for h in rows[0]]
        try:
            idx_id = headers.index("id")
            idx_name = headers.index("name")
            idx_path = headers.index("path")
            idx_params = headers.index("params")
        except ValueError as e:
            logger.error(
                "Error in Excel format: Missing column (%s)", e, exc_info=True
            )
            return
        data_list: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if not row[idx_name]:
                continue
            row_data: Dict[str, Any] = {
                "id": str(row[idx_id]) if row[idx_id] else None,
                "name": str(row[idx_name]),
                "path": str(row[idx_path]) if row[idx_path] else "",
                "params": [],
            }
            raw_params = row[idx_params]
            if raw_params:
                try:
                    row_data["params"] = json.loads(str(raw_params))
                except json.JSONDecodeError:
                    logger.warning(
                        "Warning: Invalid JSON in '%s'. Using empty list.",
                        row_data["name"],
                    )
            data_list.append(row_data)
        self._process_import_data(data_list)

    def _process_import_data(
        self, data: Union[List[Dict[str, Any]], Dict[str, Any]]
    ) -> None:
        """Common logic for inserting into the database.

        Logic: Bulk inserts/updates models."""
        if not isinstance(data, list):
            return
        with db.atomic():
            for item in data:
                if "name" in item:
                    self.add_or_update_model(
                        model_id=item.get("id"),
                        name=item["name"],
                        path=item.get("path", ""),
                        params=item.get("params", []),
                    )
